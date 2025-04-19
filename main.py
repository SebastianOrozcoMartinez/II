'''
II V0.3
This code helps identifing difrent people using UIDs (Unique Identificator Device)
and saving his assistant in a Excel file
'''

# initialize variables and import libraries
from openpyxl import load_workbook
# from adafruit_pn532.i2c import PN532_I2C
import time, re#, board, busio
from config import student, studentList
import config
wb = load_workbook("ii.xlsx")
data = wb['Hoja1']
config_sheet = wb['Ajustes']
# Start I2C
# i2c = busio.I2C(board.SCL, board.SDA)
# pn532 = PN532_I2C(i2c)

def numToCol(num):
    # This function turns any number to a excel column
    # Example: Input: 1 Output: A, Input: 4 Output: D
    Alphabet = {
        1 : 'A', 2 : 'B', 3 : 'C',
        4 : 'D', 5 : 'E', 6 : 'F',
        7 : 'G', 8 : 'H', 9 : 'I',
        10 : 'J', 11 : 'K', 12 : 'L',
        13 : 'M', 14 : 'N', 15 : 'O',
        16 : 'P', 17 : 'Q', 18 : 'R',
        19 : 'S', 20 : 'T', 21 : 'U',
        22 : 'V', 23 : 'W', 24 : 'X',
        25 : 'Y', 26 : 'Z'
    }
    x = []
    while num > 0:
        letter = Alphabet.get(num)
        x.append(letter)
        num = num - 26
    return ''.join(x)
def readCols(col, hoja):
    lista = []
    for i in range(1, 1000):
        celda = hoja.cell(row=i, column=col).value
        if celda is None:
            break
        lista.append(celda)
    return lista
def readRows(row, hoja):
    lista = []
    for i in range(1, 1000):
        celda = hoja.cell(row=row, column=i).value
        if celda is None:
            break
        lista.append(celda)
    return lista
while 1: # Attendance mode

    # Searches for a PICC
    print("Aproxima una tarjeta")
    uid = input("debug ") # "4:C0:64:B2:5B:13:90"
    '''
    while uid is None:
        uid = pn532.read_passive_target()
        if uid is not None:
            uid = "".join([hex(i) for i in uid])
            uid = regex.sub(r'0x', ':', uid).replace(":","",1).upper()
            print(uid)
            time.sleep(0.5)
    '''
    # Takes the UID and then checks if its on the list
    student_obj = studentList.get(uid)
    if student_obj is None:
        print('No se encontró el UID ⚠️')
        continue
    student = student_obj.name
    # Search for the classroom of the student
    grado = student_obj.classroom
    sheet = wb[grado]
    # Checks and compares the actual time with the check-in time 
    tiempoActual = time.localtime()
    localTime = tiempoActual.tm_hour
    if localTime >= config.entrada:
        print(f'El alumno {student} de {grado} llegó tarde')
        asistencia = False
    else:
        print(f'El alumno {student} de {grado} llegó temprano')
        asistencia = True
    # Gets the actual row
    students_por_grado = readCols(1,sheet)
    row = students_por_grado.index(student) + 1
    print(f"Actual row: {row}")
    # Get the actual column
    day = time.strftime("%d/%m/%Y")
    days = readRows(1, sheet)
    # Check if the day is already in the sheet
    if day not in days:
        col = len(days)+1
        sheet.cell(row = 1, column = len(days)+1, value = day)
        wb.save('ii.xlsx')
        days = readRows(1, sheet)
        print('El dia no estaba en la lista')
        # Here is the problem, it doesn't update the days list until the program is restarted
    else:
        col = days.index(day) + 1
    print(f"Actual col: {col}")
    # Write if the student is present or not
    if asistencia == True:
       sheet.cell(row = int(row), column = col, value = 'O')
    else:
        sheet.cell(row = int(row), column = col, value = 'X')

    # Save changes
    wb.save('ii.xlsx')
    print('Se ha registrado la asistencia')
